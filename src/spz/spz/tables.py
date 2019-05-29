# -*- coding: utf-8 -*-

"""Table export utility.

   Used to format course lists for download.
"""

import csv
import io
import re
from tempfile import NamedTemporaryFile
from openpyxl import Workbook
from openpyxl.worksheet.table import Table
from openpyxl.workbook.child import INVALID_TITLE_REGEX

from flask import make_response, url_for, redirect, flash


def export_course_list(courses, format, filename='Kursliste'):
    if format == 'csv':
        return export(CSVWriter(), courses, filename)
    elif format == 'xlsx':
        return export(ExcelWriter(), courses, filename)
    else:
        flash('Ungueltiges Export-Format: {0}'.format(format), 'error')
        return redirect(url_for('lists'))


class CSVWriter:

    mimetype = 'text/csv'
    filetype = 'csv'

    def __init__(self):
        self.buf = io.StringIO()
        self.out = csv.writer(self.buf, delimiter=";", dialect=csv.excel)
        self.header_written = False

    def write_heading(self, values):
        if not self.header_written:
            self.write_row(values)
            self.header_written = True

    def write_row(self, values):
        string_values = [str(v) if v else '' for v in values]
        self.out.writerow(string_values)

    def new_section(self, name):
        # CSV does not support sections
        pass

    def get_data(self):
        return self.buf.getvalue()


class ExcelWriter:

    mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    filetype = 'xlsx'

    def __init__(self):
        # write_only=True would require additional logic to keep track of sheet dimension so we keep it at False
        # (see sheet.dimensions in end_section())
        self.workbook = Workbook(write_only=False)
        self.workbook._sheets.clear()  # start off with no sheets

    def write_heading(self, values):
        self.write_row(values)

    def write_row(self, values):
        self.workbook._sheets[-1].append(values)

    def new_section(self, name):
        if self.workbook._sheets:
            self.end_section()
        # Sheet naming is very constrained: some chars are disallowed , names need to be unique & maximum length is 32
        # Openpyxl knows about the first two but not about the last one. It will automatically enforce unique naming by
        # appending an incrementer (e.g. [sheet, sheet] -> [sheet1, sheet2]). This then might cause the length limit to
        # be exceeded: Therefore we use a shorter limit of only 30 characters here.
        name = re.sub(INVALID_TITLE_REGEX, '', name)[:30]
        self.workbook.create_sheet(name)

    def get_data(self):
        if self.workbook._sheets:
            self.end_section()
        with NamedTemporaryFile() as file:
            self.workbook.save(file.name)
            file.seek(0)
            stream = file.read()
        return stream

    def end_section(self):
        sheet = self.workbook._sheets[-1]
        # if there are values: create a table within the excel sheet to simplify sorting by values
        # else: excel considers the file invalid if a table contains only one row so don't
        if sheet.max_row > 1:
            tableName = sheet.title.replace(' ', '_')  # needs to be unique and must not contain spaces
            table = Table(displayName=tableName, ref=sheet.dimensions)
            sheet.add_table(table)


def export(writer, courses, filename):
    # XXX: header -- not standardized
    header = ['Kurs', 'Kursplatz', 'Bewerbernummer', 'Vorname', 'Nachname', 'Mail',
              'Matrikelnummer', 'Telefon', 'Studienabschluss', 'Semester', 'Bewerberkreis']

    for course in courses:
        writer.new_section(course.full_name())
        writer.write_heading(header)

        active_no_debt = [attendance.applicant for attendance in course.attendances
                          if not attendance.waiting and (not attendance.has_to_pay or attendance.amountpaid > 0)]

        idx = 1
        for applicant in active_no_debt:
            writer.write_row([course.full_name(),
                              idx,
                              applicant.id,
                              applicant.first_name,
                              applicant.last_name,
                              applicant.mail,
                              applicant.tag,
                              applicant.phone,
                              applicant.degree.name if applicant.degree else None,
                              applicant.semester,
                              applicant.origin.name if applicant.origin else None])
            idx += 1

    resp = make_response(writer.get_data())
    resp.headers['Content-Disposition'] = 'attachment; filename="{0}.{1}"'.format(filename, writer.filetype)
    resp.mimetype = writer.mimetype

    return resp
