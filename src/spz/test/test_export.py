# -*- coding: utf-8 -*-

"""Tests the application views.
"""

from tempfile import TemporaryFile, NamedTemporaryFile
from openpyxl import load_workbook
from zipfile import ZipFile

from spz import db
from spz.export import export_course_list
from spz.models import ExportFormat, Graduation, Attendance
from test.sample_data import make_applicant


def fill(courses):
    db.session.autoflush = False
    graduation = Graduation.query.first()
    count = 0
    for course in courses:
        while not course.is_full:
            applicant = make_applicant(id=count)
            applicant.add_course_attendance(
                course=course,
                graduation=graduation,
                waiting=False,
                discount=Attendance.MAX_DISCOUNT)
            count += 1
        db.session.commit()
    return count


def test_csv(courses):
    count = fill(courses)
    format = ExportFormat.query.filter(ExportFormat.formatter == 'csv').first()
    resp = export_course_list(courses, format=format)
    assert(resp.mimetype == 'text/csv')
    assert(resp.data.count(b'\n') == count + 1)  # data + header


def test_excel(courses):
    fill(courses)
    format = ExportFormat.query.filter(ExportFormat.formatter == 'excel').first()
    resp = export_course_list(courses, format=format)
    with NamedTemporaryFile(suffix='.xlsx') as file:
        file.write(resp.data)
        wb = load_workbook(file.name)
    assert(len(wb.worksheets) == len(courses))


def test_zip_excel(courses):
    fill(courses)
    format = ExportFormat.query.filter(ExportFormat.formatter == 'zip-excel').first()
    resp = export_course_list(courses, format=format)
    assert(resp.mimetype == 'application/zip')
    with TemporaryFile() as file:
        file.write(resp.data)
        zip = ZipFile(file, 'r')
    assert(len(zip.namelist()) == len(courses))


def test_single_excel(courses):
    count = fill(courses)
    format = ExportFormat.query.filter(ExportFormat.formatter == 'single-excel').first()
    resp = export_course_list(courses, format=format)
    with NamedTemporaryFile(suffix='.xlsx') as file:
        file.write(resp.data)
        wb = load_workbook(file.name)
    assert(wb.worksheets[0].max_row - 1 >= count)  # max_row is 1 based
